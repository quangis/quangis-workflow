#!/usr/bin/env python3
"""
"""

# https://openpyxl.readthedocs.io/en/stable/usage.html#write-a-workbook
# Query Variant Options {Workflows}
# ?     ?       ?       1 1 0 1 1

from __future__ import annotations

import importlib.machinery
import importlib.util
from rdflib import Graph  # type: ignore
from transformation_algebra.query import Query
from typing import Iterator
from rdflib.term import Node
from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle

from config import query_paths, build_path, REPO  # type: ignore


def extract_queries() -> Iterator[tuple[str, str, set[Node], Query]]:
    """
    Extract queries defined in modules.
    """
    # Perhaps overengineered but makes it simple to add and change queries in
    # dedicated modules
    for path in query_paths:
        name = path.stem
        loader = importlib.machinery.SourceFileLoader(name, str(path))
        spec = importlib.util.spec_from_loader(name, loader)
        assert spec
        module = importlib.util.module_from_spec(spec)
        loader.exec_module(module)

        workflows = getattr(module, 'workflows')
        for variant in dir(module):
            query = getattr(module, variant)
            if isinstance(query, Query):
                yield name, variant, workflows, query


def result(actual: set, expected: set,
        correct=lambda x: f"\033[32m{x}\033[0m",
        wrong=lambda x: f"\033[31m{x}\033[0m",
        empty='none') -> str:
    """
    Stylize the contents of a set depending on whether its elements should be
    there or not.
    """
    if actual:
        return ", ".join(correct(r) if r in expected else wrong(r)
            for r in actual)
    else:
        return (correct if not expected else wrong)(empty)


wfgraph = Graph(store='SPARQLStore')
wfgraph.open("http://localhost:3030/name")

# Varying options to try
option_variants = {
    "ordered": {"by_types": True, "by_order": True},
    "types": {"by_types": True, "by_order": False}
}

queries = list(
    (scenario, variant, expected, query)
    for scenario, variant, expected, query in extract_queries()
    if (variant != "eval_aquifer_simon" and variant != "eval_solar")
)
all_workflows: list[Node] = list(
    set.union(*(expected for (_, _, expected, _) in queries))
)

wb = Workbook()
header = ["Scenario", "Variant", "Options"] + [
    str(wf)[len(REPO):] for wf in all_workflows]

sheet_actual = wb.active
sheet_actual.title = "Actual"
sheet_actual.append(header)

font_wrong = Font(color="FF0000")
row = 1

for scenario, variant, expected_workflows, query in queries:

    for optname, options in option_variants.items():

        query2 = Query(query.language, query.flow, **options)

        print(f'\033[1m\033[4m{scenario}\033[0m', variant, optname)

        # Print query to file for inspection
        path = build_path / f"{scenario}_{variant}_{optname}.rq"
        # print("Building", path.name)
        with open(path, 'w') as f:
            f.write(query2.sparql())

        # Run the query
        try:
            results = wfgraph.query(query2.sparql())
            positives = set(r.workflow for r in results)

            # Write to spreadsheet
            row += 1
            sheet_actual.append([scenario, variant, optname])
            for col, wf in enumerate(all_workflows, start=4):
                cell = sheet_actual.cell(row, col, "1" if wf in positives else "0")
                if (wf in positives) ^ (wf in expected_workflows):
                    cell.font = font_wrong

            # Write to terminal
            false_pos = (positives - expected_workflows)
            false_neg = (expected_workflows - positives)
            correct = positives - false_pos - false_neg

            print("Correct:", result(correct, expected_workflows))
            print("False positives:", result(false_pos, set()))
            print("Missing:", result(false_neg, set()))

        except ValueError:
            print("Not firing queries, since the server is down.")

        print()

sheet_expected = wb.create_sheet(title="Expected")
sheet_expected.append(header)
for scenario, variant, expected_workflows, _ in queries:
    for optname in option_variants.keys():
        sheet_expected.append([scenario, variant, optname] + [
            "1" if wf in expected_workflows else "0" for wf in all_workflows
        ])

wb.save(filename=build_path / "results.xlsx")
